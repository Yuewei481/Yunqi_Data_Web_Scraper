from __future__ import annotations

from copy import copy
from io import BytesIO
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter


SHEET_NAME = "贺卡"
PRODUCT_ID_HEADER = "商品ID"
IMAGE_SIZE_PX = 400


def sheet(workbook):
    return workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]


def copy_cell_style(source, target):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def find_header_col(ws, header):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value and str(value).strip() == header:
            return col
    raise RuntimeError(f"表格缺少表头: {header}")


def product_rows(ws):
    product_id_col = find_header_col(ws, PRODUCT_ID_HEADER)
    rows = {}
    for row in range(2, ws.max_row + 1):
        product_id = ws.cell(row, product_id_col).value
        if product_id is None:
            continue
        product_id = str(product_id).strip()
        if product_id:
            rows[product_id] = row
    return rows


def export_product_ids(excel_path, output_path):
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    ws = sheet(workbook)
    ids = list(product_rows(ws).keys())
    workbook.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(ids, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已导出商品ID: {len(ids)}")
    print(output_path)


def last_product_row(ws):
    rows = product_rows(ws)
    return max(rows.values()) if rows else 1


def image_bytes(image):
    data = image._data()
    return data if isinstance(data, bytes) else bytes(data)


def images_by_row(ws, needed_rows):
    needed_rows = set(needed_rows)
    result = {}
    for image in ws._images:
        row = image.anchor._from.row + 1
        if row not in needed_rows:
            continue
        col = image.anchor._from.col + 1
        result.setdefault(row, []).append(
            {
                "col": col,
                "bytes": image_bytes(image),
            }
        )
    return result


def ensure_columns(target_ws, source_ws):
    if target_ws.max_column >= source_ws.max_column:
        return
    start_col = target_ws.max_column + 1
    amount = source_ws.max_column - target_ws.max_column
    target_ws.insert_cols(start_col, amount)
    for col in range(start_col, source_ws.max_column + 1):
        source_cell = source_ws.cell(1, col)
        target_cell = target_ws.cell(1, col)
        copy_cell_style(source_cell, target_cell)
        target_cell.value = source_cell.value
        source_dim = source_ws.column_dimensions[get_column_letter(col)]
        target_dim = target_ws.column_dimensions[get_column_letter(col)]
        if source_dim.width:
            target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden


def copy_row(source_ws, target_ws, source_row, target_row, source_images):
    for col in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(source_row, col)
        target_cell = target_ws.cell(target_row, col)
        copy_cell_style(source_cell, target_cell)
        target_cell.value = source_cell.value

    source_height = source_ws.row_dimensions[source_row].height
    if source_height:
        target_ws.row_dimensions[target_row].height = source_height

    for item in source_images.get(source_row, []):
        img = ExcelImage(BytesIO(item["bytes"]))
        img.width = IMAGE_SIZE_PX
        img.height = IMAGE_SIZE_PX
        target_ws.add_image(img, target_ws.cell(target_row, item["col"]).coordinate)


def update_auto_filter(ws):
    if ws.auto_filter:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"


def main():
    if len(sys.argv) == 4 and sys.argv[1] == "--export-ids":
        export_product_ids(Path(sys.argv[2]).expanduser().resolve(), Path(sys.argv[3]).expanduser().resolve())
        return

    if len(sys.argv) != 3:
        raise SystemExit("Usage: append_yunqi_new_products.py base.xlsx today.xlsx\n       append_yunqi_new_products.py --export-ids base.xlsx ids.json")

    base_path = Path(sys.argv[1]).expanduser().resolve()
    today_path = Path(sys.argv[2]).expanduser().resolve()

    base_wb = load_workbook(base_path)
    today_wb = load_workbook(today_path)
    base_ws = sheet(base_wb)
    today_ws = sheet(today_wb)

    base_rows = product_rows(base_ws)
    today_rows = product_rows(today_ws)
    new_ids = sorted(set(today_rows) - set(base_rows), key=lambda product_id: today_rows[product_id])

    ensure_columns(base_ws, today_ws)

    today_images = images_by_row(today_ws, [today_rows[product_id] for product_id in new_ids])
    target_row = last_product_row(base_ws) + 1
    for product_id in new_ids:
        copy_row(today_ws, base_ws, today_rows[product_id], target_row, today_images)
        target_row += 1

    update_auto_filter(base_ws)
    base_wb.save(base_path)
    today_wb.close()
    base_wb.close()

    print(f"新增商品: {len(new_ids)}")
    print(base_path)


if __name__ == "__main__":
    main()
