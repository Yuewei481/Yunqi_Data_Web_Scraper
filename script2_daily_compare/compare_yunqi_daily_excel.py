from __future__ import annotations

from copy import copy
from io import BytesIO
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAME = "贺卡"
PRODUCT_ID_HEADER = "商品ID"
STATUS_HEADER = "状态"
STATUS_ADDED = "今日新增"
STATUS_REMOVED = "今日移除"
IMAGE_SIZE_PX = 400
ADDED_FILL = PatternFill("solid", fgColor="FFF2CC")
REMOVED_FILL = PatternFill("solid", fgColor="D9EAF7")


def sheet(workbook):
    return workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]


def copy_cell_style(source, target):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def find_header_col(ws, header):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value and str(value).strip() == header:
            return col
    raise RuntimeError(f"表格缺少表头: {header}")


def product_rows(ws):
    product_id_col = find_header_col(ws, PRODUCT_ID_HEADER)
    rows = {}
    for row in range(2, ws.max_row + 1):
        product_id = ws.cell(row, product_id_col).value
        if product_id is None:
            continue
        product_id = str(product_id).strip()
        if product_id:
            rows[product_id] = row
    return rows


def read_product_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = sheet(workbook)
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    product_id_col = None
    for index, value in enumerate(header_values, start=1):
        if value and str(value).strip() == PRODUCT_ID_HEADER:
            product_id_col = index
            break
    if product_id_col is None:
        workbook.close()
        raise RuntimeError(f"{path} 缺少表头: {PRODUCT_ID_HEADER}")

    rows = {}
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if product_id_col > len(row):
            continue
        product_id = row[product_id_col - 1]
        if product_id is None:
            continue
        product_id = str(product_id).strip()
        if product_id:
            rows[product_id] = row_index
    max_column = ws.max_column
    workbook.close()
    return rows, max_column


def clear_output_rows(ws, start_row=2):
    ws._images = []
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def image_bytes(image):
    data = image._data()
    return data if isinstance(data, bytes) else bytes(data)


def images_by_row(ws, needed_rows):
    needed_rows = set(needed_rows)
    result = {}
    for image in ws._images:
        row = image.anchor._from.row + 1
        if row not in needed_rows:
            continue
        col = image.anchor._from.col + 1
        result.setdefault(row, []).append(
            {
                "col": col,
                "bytes": image_bytes(image),
                "width": image.width,
                "height": image.height,
            }
        )
    return result


def copy_column_dimensions(source_ws, output_ws, max_col):
    output_ws.column_dimensions["A"].width = 14
    for col in range(1, max_col + 1):
        source_letter = get_column_letter(col)
        target_letter = get_column_letter(col + 1)
        source_dim = source_ws.column_dimensions[source_letter]
        target_dim = output_ws.column_dimensions[target_letter]
        if source_dim.width:
            target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden


def copy_row(source_ws, output_ws, source_row, target_row, status, source_images, max_source_col):
    status_cell = output_ws.cell(target_row, 1)
    copy_cell_style(output_ws.cell(1, 1), status_cell)
    status_cell.value = status
    status_cell.fill = ADDED_FILL if status == STATUS_ADDED else REMOVED_FILL

    for col in range(1, max_source_col + 1):
        source_cell = source_ws.cell(source_row, col)
        target_cell = output_ws.cell(target_row, col + 1)
        copy_cell_style(source_cell, target_cell)
        target_cell.value = source_cell.value

    source_height = source_ws.row_dimensions[source_row].height
    if source_height:
        output_ws.row_dimensions[target_row].height = source_height

    for item in source_images.get(source_row, []):
        img = ExcelImage(BytesIO(item["bytes"]))
        img.width = IMAGE_SIZE_PX
        img.height = IMAGE_SIZE_PX
        output_ws.add_image(img, output_ws.cell(target_row, item["col"] + 1).coordinate)


def prepare_output(today_ws, max_source_col):
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = today_ws.title

    copy_cell_style(today_ws.cell(1, 1), output_ws.cell(1, 1))
    output_ws.cell(1, 1).value = STATUS_HEADER

    for col in range(1, max_source_col + 1):
        source_cell = today_ws.cell(1, col)
        target_cell = output_ws.cell(1, col + 1)
        copy_cell_style(source_cell, target_cell)
        target_cell.value = source_cell.value

    if today_ws.row_dimensions[1].height:
        output_ws.row_dimensions[1].height = today_ws.row_dimensions[1].height
    output_ws.freeze_panes = "A2"
    output_ws.auto_filter.ref = f"A1:{get_column_letter(max_source_col + 1)}1"
    return output_wb, output_ws


def main():
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: compare_yunqi_daily_excel.py yesterday.xlsx today.xlsx diff_output.xlsx"
        )

    yesterday_path = Path(sys.argv[1]).expanduser().resolve()
    today_path = Path(sys.argv[2]).expanduser().resolve()
    output_path = Path(sys.argv[3]).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    yesterday_rows, yesterday_max_col = read_product_rows(yesterday_path)
    today_rows, today_max_col = read_product_rows(today_path)
    yesterday_ids = set(yesterday_rows)
    today_ids = set(today_rows)

    added_ids = sorted(today_ids - yesterday_ids, key=lambda pid: today_rows[pid])
    removed_ids = sorted(yesterday_ids - today_ids, key=lambda pid: yesterday_rows[pid])

    max_source_col = max(yesterday_max_col, today_max_col)
    today_wb = load_workbook(today_path)
    today_ws = sheet(today_wb)
    output_wb, output_ws = prepare_output(today_ws, max_source_col)
    copy_column_dimensions(today_ws, output_ws, max_source_col)

    yesterday_wb = None
    yesterday_ws = None
    if removed_ids:
        yesterday_wb = load_workbook(yesterday_path)
        yesterday_ws = sheet(yesterday_wb)

    today_needed_rows = [today_rows[product_id] for product_id in added_ids]
    yesterday_needed_rows = [yesterday_rows[product_id] for product_id in removed_ids]
    today_images = images_by_row(today_ws, today_needed_rows)
    yesterday_images = images_by_row(yesterday_ws, yesterday_needed_rows) if yesterday_ws else {}

    target_row = 2
    for product_id in added_ids:
        copy_row(today_ws, output_ws, today_rows[product_id], target_row, STATUS_ADDED, today_images, max_source_col)
        target_row += 1
    for product_id in removed_ids:
        copy_row(yesterday_ws, output_ws, yesterday_rows[product_id], target_row, STATUS_REMOVED, yesterday_images, max_source_col)
        target_row += 1

    output_wb.save(output_path)
    today_wb.close()
    if yesterday_wb:
        yesterday_wb.close()
    print(f"今日新增: {len(added_ids)}")
    print(f"今日移除: {len(removed_ids)}")
    print(output_path)


if __name__ == "__main__":
    main()
