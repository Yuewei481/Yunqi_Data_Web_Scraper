from __future__ import annotations

from copy import copy
import json
import shutil
import sys
from pathlib import Path

from openpyxl.packaging import manifest
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter


IMAGE_SIZE_PX = 400


REQUIRED_HEADERS = {
    "listedAt": "上架时间",
    "imagePath": "参考样品图",
    "dailySales": "日销",
    "monthlySales": "月销",
    "priceUsd": "售价",
    "title": "标题",
    "productId": "商品ID",
}


EXTRA_IMAGE_HEADER = "竞品内容图片"
THEME_HEADER = "主题"
PRODUCT_ID_HEADER = "商品ID"


def find_columns(ws):
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value:
            headers[str(value).strip()] = col

    missing = [header for header in REQUIRED_HEADERS.values() if header not in headers]
    if missing:
        raise RuntimeError(f"模板缺少这些表头: {', '.join(missing)}")
    return {key: headers[name] for key, name in REQUIRED_HEADERS.items()}


def find_header_col(ws, header):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value and str(value).strip() == header:
            return col
    return None


def clear_old_rows(ws, start_row=2):
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def column_width_for_pixels(px):
    return min(255, max(16, px / 7))


def row_height_for_pixels(px):
    return max(78, px * 0.75)


def copy_cell_style(source, target):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def ensure_extra_image_columns(ws, start_col, extra_image_count):
    if extra_image_count <= 0:
        return
    header_template = ws.cell(1, start_col)
    if extra_image_count > 1:
        ws.insert_cols(start_col + 1, amount=extra_image_count - 1)
    for offset in range(extra_image_count):
        header_cell = ws.cell(1, start_col + offset)
        copy_cell_style(header_template, header_cell)
        header_cell.value = f"额外参考样品图{offset + 1}"
    ws.auto_filter.ref = f"A1:{get_column_letter(start_col + extra_image_count - 1)}1"


def prepare_columns(ws):
    daily_col = find_header_col(ws, "日销")
    if not daily_col:
        raise RuntimeError("模板缺少表头: 日销")

    product_id_col = find_header_col(ws, PRODUCT_ID_HEADER)
    if not product_id_col:
        extra_col = find_header_col(ws, EXTRA_IMAGE_HEADER)
        title_col = find_header_col(ws, "标题")
        insert_col = extra_col or ((title_col or ws.max_column) + 1)
        ws.insert_cols(insert_col, 1)
        product_id_col = insert_col
    product_id_cell = ws.cell(1, product_id_col)
    copy_cell_style(ws.cell(1, daily_col), product_id_cell)
    product_id_cell.value = PRODUCT_ID_HEADER

    theme_col = find_header_col(ws, THEME_HEADER)
    if theme_col:
        ws.delete_cols(theme_col, 1)


def set_image_column_sizes(ws, image_cols):
    width = column_width_for_pixels(IMAGE_SIZE_PX)
    for col in image_cols:
        ws.column_dimensions[get_column_letter(col)].width = width


def add_images(ws, image_paths, image_cols, row):
    paths = [Path(p) for p in image_paths or [] if p]
    paths = [p for p in paths if p.exists()]
    if not paths:
        return

    for image_path, col in zip(paths, image_cols):
        img = ExcelImage(str(image_path))
        img.width = IMAGE_SIZE_PX
        img.height = IMAGE_SIZE_PX
        ws.add_image(img, ws.cell(row, col).coordinate)

    ws.row_dimensions[row].height = row_height_for_pixels(IMAGE_SIZE_PX)


def add_image(ws, image_path, cell, row):
    if not image_path:
        return
    p = Path(image_path)
    if not p.exists():
        return
    img = ExcelImage(str(p))
    img.width = IMAGE_SIZE_PX
    img.height = IMAGE_SIZE_PX
    ws.add_image(img, cell)
    ws.row_dimensions[row].height = row_height_for_pixels(IMAGE_SIZE_PX)


def main():
    manifest.mimetypes.add_type("image/webp", ".webp")

    if len(sys.argv) != 4:
        raise SystemExit("Usage: fill_yunqi_excel.py products.json template.xlsx output.xlsx")

    data_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    products = json.loads(data_path.read_text(encoding="utf-8"))
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb["贺卡"] if "贺卡" in wb.sheetnames else wb[wb.sheetnames[0]]
    prepare_columns(ws)
    initial_cols = find_columns(ws)
    max_images = max([len(product.get("imagePaths") or []) for product in products] or [1])
    max_extra_images = max(0, max_images - 1)
    extra_start_col = find_header_col(ws, EXTRA_IMAGE_HEADER) or initial_cols["productId"] + 1
    ensure_extra_image_columns(ws, extra_start_col, max_extra_images)
    cols = find_columns(ws)
    clear_old_rows(ws)

    extra_image_cols = [extra_start_col + offset for offset in range(max_extra_images)]
    set_image_column_sizes(ws, [cols["imagePath"], *extra_image_cols])
    ws.column_dimensions[ws.cell(1, cols["title"]).column_letter].width = 56
    ws.column_dimensions[ws.cell(1, cols["productId"]).column_letter].width = 24
    ws.column_dimensions[ws.cell(1, cols["listedAt"]).column_letter].width = 16

    for idx, product in enumerate(products, start=2):
        ws.cell(idx, cols["listedAt"], product.get("listedAt", ""))
        ws.cell(idx, cols["dailySales"], product.get("dailySales", ""))
        ws.cell(idx, cols["monthlySales"], product.get("monthlySales", ""))
        ws.cell(idx, cols["priceUsd"], product.get("priceUsd", ""))
        ws.cell(idx, cols["title"], product.get("title", ""))
        ws.cell(idx, cols["productId"], product.get("productId", ""))

        image_paths = product.get("imagePaths") or []
        if image_paths:
            add_images(ws, image_paths[:1], [cols["imagePath"]], idx)
            add_images(ws, image_paths[1:], extra_image_cols, idx)
        else:
            image_cell = ws.cell(idx, cols["imagePath"]).coordinate
            add_image(ws, product.get("imagePath", ""), image_cell, idx)

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
